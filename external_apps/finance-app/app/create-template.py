#!/usr/bin/env python3
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Create a new workbook
wb = Workbook()
ws = wb.active
ws.title = "REALISASI"

# Add header rows
data = [
    ['LAPORAN REALISASI ANGGARAN'],
    [''],
    ['TABEL 1 - SEKSI PENGEMBANGAN'],
    [''],
    ['BAGIAN', ':', 'SEKSI PENGEMBANGAN'],
    ['SUMBER ANGGARAN', ':', 'PAD MURNI'],
    ['PROGRAM', ':', 'PROGRAM PENUNJANG URUSAN PEMERINTAHAN DAERAH PROVINSI'],
    ['Kegiatan', ':', ''],
    ['Sub Kegiatan', ':', ''],
    [''],
    ['NO', 'KODE REKENING', 'URAIAN', 'SEBELUM', 'SETELAH', 'JUMLAH', 'PAPBD', 'JAN', 'FEB', 'MAR', 'APR', 'MEI', 'JUN', 'JUL', 'AGU', 'SEP', 'OKT', 'NOV', 'DES', 'JUMLAH REALISASI'],
]

# Write data to worksheet
for row_idx, row_data in enumerate(data, start=1):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Bold header row
        if row_idx == 11:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

# Set column widths
column_widths = [5, 15, 40, 12, 12, 12, 12, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 12]
for idx, width in enumerate(column_widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

# Save the workbook
wb.save('public/DKKB DAN REALISASI.xlsx')
print('✅ Template Excel file created successfully at: public/DKKB DAN REALISASI.xlsx')
