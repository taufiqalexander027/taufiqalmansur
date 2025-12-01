from openpyxl import load_workbook

# Load workbook
wb = load_workbook('DKKB DAN REALISASI.xlsx', data_only=True)

print("=" * 100)
print("ANALISIS DETAIL SHEET: DKKB 2025")
print("=" * 100)

ws_dkkb = wb['DKKB 2025']

# Show first 30 rows
print("\n� 50 BARIS PERTAMA:\n")
for row in range(1, 51):
    row_data = []
    has_data = False
    for col in range(1, 20):  # First 20 columns
        value = ws_dkkb.cell(row, col).value
        if value is not None:
            has_data = True
        if value and isinstance(value, str) and len(value) > 50:
            value = value[:50] + "..."
        row_data.append(value)
    
    if has_data:
        print(f"Baris {row:3d}: {row_data}")

print("\n" + "=" * 100)
